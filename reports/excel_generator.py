"""
Gerador de Relatórios Excel - Sistema Zeca Delivery
==================================================

Este módulo consome a API de entregas e gera relatórios Excel formatados
com cores condicionais e estatísticas automáticas.

Funcionalidades:
- Conecta com a API Flask
- Gera Excel com duas abas: Entregas + Estatísticas
- Formatação profissional com cores por status
- Ajuste automático de colunas
- Tratamento de erros robusto

Autor: Demonstração do artigo Zeca Delivery
"""

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os

class DeliveryReportGenerator:
    """Classe para gerar relatórios Excel das entregas"""
    
    def __init__(self, api_url="http://localhost:5000"):
        self.api_url = api_url
        
    def verify_api_connection(self):
        """Verifica se a API está online e funcionando"""
        try:
            print("🔍 Verificando conexão com a API...")
            response = requests.get(f"{self.api_url}/api/health", timeout=5)
            
            if response.status_code == 200:
                health_data = response.json()
                service_name = health_data.get('service', 'API')
                print(f"✅ API está online: {service_name}")
                return True
            else:
                print(f"❌ API retornou status {response.status_code}")
                return False
                
        except requests.exceptions.ConnectionError:
            print("❌ Erro: Não foi possível conectar com a API")
            print("💡 Certifique-se de que a API está rodando em http://localhost:5000")
            return False
        except Exception as e:
            print(f"❌ Erro ao conectar com a API: {e}")
            return False
    
    def fetch_deliveries(self):
        """Busca dados das entregas da API"""
        print("📡 Buscando dados das entregas...")
        try:
            response = requests.get(f"{self.api_url}/api/entregas", timeout=10)
            response.raise_for_status()
            
            data = response.json()
            deliveries = data.get('data', [])
            print(f"✅ {len(deliveries)} entregas encontradas")
            return deliveries
            
        except requests.exceptions.RequestException as e:
            print(f"❌ Erro ao buscar entregas: {e}")
            return []
        except Exception as e:
            print(f"❌ Erro inesperado: {e}")
            return []
    
    def fetch_statistics(self):
        """Busca estatísticas da API"""
        try:
            response = requests.get(f"{self.api_url}/api/stats", timeout=10)
            response.raise_for_status()
            
            data = response.json()
            stats = data.get('data', {})
            return stats
            
        except Exception as e:
            print(f"⚠️ Aviso: Não foi possível buscar estatísticas: {e}")
            return {}
    
    def create_styled_workbook(self, deliveries):
        """Cria planilha Excel com formatação profissional"""
        print("📊 Criando planilha Excel com formatação...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Entregas do Dia"
        
        # Definir cabeçalhos
        headers = [
            "ID", "Cliente", "Endereço Completo", "Produto", 
            "Quantidade", "Valor (R$)", "Status", "Entrega Prevista", 
            "Telefone", "Prioridade"
        ]
        
        # Inserir cabeçalhos
        ws.append(headers)
        
        # Estilo dos cabeçalhos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar estilo aos cabeçalhos
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Inserir dados das entregas
        for row_num, delivery in enumerate(deliveries, 2):
            # Formatar endereço completo
            endereco_completo = f"{delivery['endereco']}, {delivery['bairro']}, {delivery['cidade']} - {delivery['cep']}"
            
            # Formatar data de entrega
            entrega_formatada = delivery['entrega_prevista'].replace('T', ' ')
            
            # Dados da linha
            row_data = [
                delivery['id'],
                delivery['cliente'],
                endereco_completo,
                delivery['produto'],
                delivery['quantidade'],
                delivery['valor'],
                delivery['status'].replace('_', ' ').title(),
                entrega_formatada,
                delivery['telefone'],
                delivery['prioridade'].title()
            ]
            
            ws.append(row_data)
            
            # Aplicar formatação condicional por status
            self._apply_status_formatting(ws, row_num, delivery['status'], len(headers))
        
        # Ajustar largura das colunas
        self._adjust_column_widths(ws)
        
        # Adicionar bordas
        self._add_borders(ws, len(deliveries) + 1, len(headers))
        
        return wb
    
    def _apply_status_formatting(self, ws, row_num, status, num_cols):
        """Aplica cores condicionais baseadas no status"""
        status_colors = {
            'pendente': 'FFE4B5',      # Bege claro
            'em_transito': 'E6F3FF',   # Azul claro
            'entregue': 'E8F5E8',      # Verde claro
            'cancelado': 'FFE4E1'      # Vermelho claro
        }
        
        if status in status_colors:
            fill = PatternFill(
                start_color=status_colors[status],
                end_color=status_colors[status], 
                fill_type="solid"
            )
            
            # Aplicar cor a toda a linha
            for col in range(1, num_cols + 1):
                ws.cell(row=row_num, column=col).fill = fill
    
    def _adjust_column_widths(self, ws):
        """Ajusta largura das colunas automaticamente"""
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            # Limitar largura máxima e mínima
            adjusted_width = min(max(length + 2, 12), 50)
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    
    def _add_borders(self, ws, num_rows, num_cols):
        """Adiciona bordas à tabela"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, num_rows + 1):
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = thin_border
    
    def add_statistics_sheet(self, wb, statistics):
        """Adiciona aba com estatísticas"""
        if not statistics:
            return
            
        print("📈 Adicionando aba de estatísticas...")
        
        ws_stats = wb.create_sheet("Estatísticas")
        
        # Título
        ws_stats['A1'] = "📊 Estatísticas das Entregas"
        ws_stats['A1'].font = Font(bold=True, size=14, color="366092")
        
        # Dados estatísticos
        stats_data = [
            ["📦 Total de Entregas:", statistics.get('total_entregas', 0)],
            ["💰 Valor Total:", f"R$ {statistics.get('valor_total', 0):.2f}"],
            ["📈 Taxa de Entrega:", f"{statistics.get('taxa_entrega', 0):.1f}%"],
            ["💸 Valor Médio:", f"R$ {statistics.get('valor_medio', 0):.2f}"],
            ["", ""],
            ["📋 Distribuição por Status:", ""]
        ]
        
        # Adicionar distribuição por status
        distribuicao = statistics.get('distribuicao_status', {})
        for status, count in distribuicao.items():
            stats_data.append([f"  {status.replace('_', ' ').title()}:", count])
        
        # Inserir dados
        for row_num, (label, value) in enumerate(stats_data, 3):
            ws_stats.cell(row=row_num, column=1, value=label)
            ws_stats.cell(row=row_num, column=2, value=value)
            
            # Estilo para labels
            if label and not label.startswith("  "):
                ws_stats.cell(row=row_num, column=1).font = Font(bold=True)
        
        # Ajustar larguras
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 15
    
    def generate_report(self):
        """Gera o relatório completo"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relatorio_entregas_{timestamp}.xlsx"
        
        print("🚀 Iniciando geração do relatório de entregas...")
        print("=" * 60)
        
        # Verificar conexão com API
        if not self.verify_api_connection():
            print("❌ Falha na conexão com a API. Abortando...")
            return False
        
        # Buscar dados
        deliveries = self.fetch_deliveries()
        if not deliveries:
            print("❌ Nenhuma entrega encontrada!")
            return False
        
        # Buscar estatísticas
        statistics = self.fetch_statistics()
        
        # Criar planilha
        wb = self.create_styled_workbook(deliveries)
        print("✅ Planilha criada com formatação profissional")
        
        # Adicionar estatísticas
        if statistics:
            self.add_statistics_sheet(wb, statistics)
            print("✅ Aba de estatísticas adicionada")
        
        # Salvar arquivo
        try:
            wb.save(filename)
            print(f"✅ Planilha salva como: {filename}")
        except Exception as e:
            print(f"❌ Erro ao salvar planilha: {e}")
            return False
        
        # Mostrar resumo
        self._show_summary(filename, deliveries, statistics)
        
        return True
    
    def _show_summary(self, filename, deliveries, statistics):
        """Mostra resumo final do relatório gerado"""
        print("=" * 60)
        print("🎉 Relatório gerado com sucesso!")
        print(f"📁 Arquivo: {filename}")
        print(f"📊 Total de entregas: {len(deliveries)}")
        
        if statistics:
            print(f"💰 Valor total: R$ {statistics.get('valor_total', 0):.2f}")
            print(f"📈 Taxa de entrega: {statistics.get('taxa_entrega', 0):.1f}%")
        
        print("💡 Dica: Abra o arquivo Excel para visualizar o relatório!")
        print("=" * 60)

def main():
    """Função principal para execução standalone"""
    print("📋 Gerador de Relatórios Excel - Sistema Zeca Delivery")
    print("=" * 60)
    
    generator = DeliveryReportGenerator()
    success = generator.generate_report()
    
    if not success:
        print("\n❌ Falha na geração do relatório")
        print("💡 Verifique se a API está rodando: python api/delivery_api.py")
        return 1
    
    return 0

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
