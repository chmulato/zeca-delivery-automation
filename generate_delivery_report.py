# Gerador de Relatório Excel - Sistema Zeca Delivery
# Arquivo: generate_delivery_report.py

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

class DeliveryReportGenerator:
    def __init__(self, api_url="http://localhost:5000"):
        self.api_url = api_url
        
    def verify_api_connection(self):
        """Verifica se a API está online"""
        try:
            response = requests.get(f"{self.api_url}/api/health", timeout=5)
            if response.status_code == 200:
                health_data = response.json()
                print(f"✅ API está online: {health_data.get('service', 'API')}")
                return True
            return False
        except Exception as e:
            print(f"❌ Erro ao conectar com a API: {e}")
            return False
    
    def fetch_deliveries(self):
        """Busca dados das entregas da API"""
        print("📡 Buscando dados das entregas...")
        try:
            response = requests.get(f"{self.api_url}/api/entregas")
            response.raise_for_status()
            data = response.json()
            deliveries = data.get('data', [])
            print(f"✅ {len(deliveries)} entregas encontradas")
            return deliveries
        except Exception as e:
            print(f"❌ Erro ao buscar entregas: {e}")
            return []
    
    def create_styled_workbook(self, deliveries):
        """Cria planilha Excel com formatação profissional"""
        print("📊 Criando planilha Excel...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Entregas do Dia"
        
        # Cabeçalhos
        headers = ["ID", "Cliente", "Endereço Completo", "Produto", 
                  "Valor (R$)", "Status", "Entrega Prevista", "Telefone"]
        ws.append(headers)
        
        # Estilo dos cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", 
                                end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Inserir dados com formatação condicional
        for row_num, delivery in enumerate(deliveries, 2):
            endereco_completo = f"{delivery['endereco']}, {delivery['bairro']}, {delivery['cidade']} - {delivery['cep']}"
            
            ws.append([
                delivery['id'],
                delivery['cliente'],
                endereco_completo,
                delivery['produto'],
                delivery['valor'],
                delivery['status'],
                delivery['entrega_prevista'],
                delivery['telefone']
            ])
            
            # Formatação condicional por status
            status_colors = {
                'pendente': 'FFE4B5',    # Bege claro
                'em_transito': 'E6F3FF',  # Azul claro
                'entregue': 'E8F5E8'     # Verde claro
            }
            
            status = delivery['status']
            if status in status_colors:
                fill = PatternFill(start_color=status_colors[status],
                                 end_color=status_colors[status], 
                                 fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = fill
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    def add_statistics_sheet(self, wb, deliveries):
        """Adiciona aba de estatísticas"""
        print("📈 Adicionando aba de estatísticas...")
        ws_stats = wb.create_sheet(title="Estatísticas")
        
        # Calcular estatísticas
        total = len(deliveries)
        pendentes = len([d for d in deliveries if d['status'] == 'pendente'])
        em_transito = len([d for d in deliveries if d['status'] == 'em_transito'])
        entregues = len([d for d in deliveries if d['status'] == 'entregue'])
        valor_total = sum(d['valor'] for d in deliveries)
        taxa_entrega = (entregues / total) * 100 if total > 0 else 0
        
        # Dados das estatísticas
        stats_data = [
            ["Métrica", "Valor"],
            ["Total de Entregas", total],
            ["Entregas Pendentes", pendentes],
            ["Entregas em Trânsito", em_transito],
            ["Entregas Entregues", entregues],
            ["Valor Total (R$)", f"R$ {valor_total:.2f}"],
            ["Taxa de Entrega (%)", f"{taxa_entrega:.1f}%"],
        ]
        
        # Inserir dados
        for row_data in stats_data:
            ws_stats.append(row_data)
        
        # Formatação dos cabeçalhos
        for col in range(1, 3):
            cell = ws_stats.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar largura das colunas
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 20
    
    def generate_report(self):
        """Gera o relatório completo"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relatorio_entregas_{timestamp}.xlsx"
        
        print("🚀 Iniciando geração do relatório de entregas...")
        print("=" * 50)
        
        # Verificar conexão com API
        if not self.verify_api_connection():
            return False
        
        # Buscar dados
        deliveries = self.fetch_deliveries()
        if not deliveries:
            print("❌ Nenhuma entrega encontrada!")
            return False
        
        # Criar planilha
        wb = self.create_styled_workbook(deliveries)
        print("✅ Planilha criada com formatação profissional")
        
        # Adicionar aba de estatísticas
        self.add_statistics_sheet(wb, deliveries)
        
        # Salvar arquivo
        wb.save(filename)
        print(f"✅ Planilha salva como: {filename}")
        
        # Mostrar estatísticas finais
        total_valor = sum(d['valor'] for d in deliveries)
        entregues = len([d for d in deliveries if d['status'] == 'entregue'])
        taxa_entrega = (entregues / len(deliveries)) * 100
        
        print("=" * 50)
        print("🎉 Relatório gerado com sucesso!")
        print(f"📁 Arquivo: {filename}")
        print(f"📊 Total de entregas: {len(deliveries)}")
        print(f"💰 Valor total: R$ {total_valor:.2f}")
        print(f"📈 Taxa de entrega: {taxa_entrega:.1f}%")
        print("💡 Dica: Abra o arquivo Excel para visualizar o relatório!")
        
        return True

# Execução principal
if __name__ == "__main__":
    generator = DeliveryReportGenerator()
    generator.generate_report()
